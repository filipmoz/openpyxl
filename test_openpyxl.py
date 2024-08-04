"""Basic openpyxl exporter tests."""
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

# Same real input as create_sample.py
SAMPLE_ROWS = [
    ["Product", "Qty", "Price"],
    ["Coffee", 2, 3.50],
    ["Tea", 1, 2.00],
    ["Milk", 3, 1.20],
]


def test_create_workbook():
    wb = Workbook()
    assert wb is not None
    ws = wb.active
    assert ws is not None


def test_write_read_cell():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 42
    assert ws["A1"].value == 42


def test_cell_row_column():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=3, value="hello")
    assert ws.cell(row=2, column=3).value == "hello"


def test_append_row():
    wb = Workbook()
    ws = wb.active
    ws.append([1, 2, 3])
    assert ws["A1"].value == 1
    assert ws["B1"].value == 2
    assert ws["C1"].value == 3


def test_save_load_roundtrip():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "test"
    ws["B2"] = 100
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    try:
        wb.save(path)
        wb2 = load_workbook(path)
        ws2 = wb2.active
        assert ws2["A1"].value == "test"
        assert ws2["B2"].value == 100
    finally:
        path.unlink(missing_ok=True)


def test_sheet_by_name():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = 99
    assert wb["Summary"]["A1"].value == 99
    assert wb["Data"] is ws1


def test_real_data_roundtrip():
    wb = Workbook()
    ws = wb.active
    for row in SAMPLE_ROWS:
        ws.append(row)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    try:
        wb.save(path)
        wb2 = load_workbook(path)
        ws2 = wb2.active
        for r, row in enumerate(SAMPLE_ROWS, start=1):
            for c, val in enumerate(row, start=1):
                assert ws2.cell(row=r, column=c).value == val
    finally:
        path.unlink(missing_ok=True)
